from importlib.metadata import version, PackageNotFoundError
import argparse
import csv
import re
import sys
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

# Mapping placeholders in templates to actual column names
PLACEHOLDER_MAP = {
    "HMI_TAG": "HMI_TAG",
    "PLC_TAG": "PLC_TAG",
    "COMMENT001": "COMMENT",
    "ACCESS_NAME": "ACCESS_NAME",
    "ALARM_GROUP": "ALARM_GROUP",
}

# Set of required columns that must be present in the input Excel sheet
REQUIRED_COLUMNS = {
    "HMI_TAG",
    "ACCESS_NAME",
    "DEVICE_TYPE",
}

# List of optional columns for the input Excel sheet. Ordered for consistent warning output.
OPTIONAL_COLUMNS = [
    "PLC_TAG",
    "COMMENT",
    "ALARM_GROUP",
    "OFFSET",
]

# get version number from pyproject.toml


def get_version():
    try:
        return version("wwtags")
    except PackageNotFoundError:
        return "unknown"


def parse_args():
    """
    Parse command-line arguments.

    Returns:
        argparse.Namespace: Parsed arguments.
    """

    pkg_version = get_version()

    parser = argparse.ArgumentParser(
        description=f"Generate Wonderware tag import CSV from Excel templates\nVersion {pkg_version}"
    )
    parser.add_argument(
        "workbook", help="Excel workbook containing DEVICE_LIST and templates")
    parser.add_argument(
        "--output",
        default="ww_tag_import.csv",
        help="Output CSV file (default: ww_tag_import.csv)",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Fail immediately on missing templates or required fields",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Validate and count tags without writing output file",
    )
    parser.add_argument(
        "--version",
        action="version",
        version=f"%(prog)s {pkg_version}"
    )
    parser.add_argument(
        "--list-templates",
        action="store_true",
        help="List available template sheets in the workbook and exit"
    )
    parser.add_argument(
        "--filter",
        metavar="COL=VAL",
        help="Only process rows where COL equals VAL (e.g. --filter DEVICE_TYPE=VFD)",
    )
    parser.add_argument(
        "--list-columns",
        action="store_true",
        help="List DEVICE_LIST columns (required and optional) and exit"
    )
    parser.add_argument(
        "--import-udt",
        metavar="L5X_FILE",
        help="Import a Studio5000 UDT .L5X file and add a template sheet to the workbook",
    )
    parser.add_argument(
        "--tag-length",
        action="store_true",
        help="Check the maximum HMI_TAG value length allowed by each template sheet and exit",
    )
    parser.add_argument(
        "--ww-version",
        choices=["2020", "2023"],
        help="Wonderware version for --tag-length (2020=32-char limit, 2023=128-char limit); skips interactive prompt",
    )

    return parser.parse_args()


def list_templates(wb):
    """
    Return a list of template sheet names (all sheets except DEVICE_LIST and internal $-prefixed sheets).
    """
    return [name for name in wb.sheetnames if name != "DEVICE_LIST" and not name.startswith("$")]


def read_device_list(ws, strict=False, warnings=None, errors=None):
    """
    Read the DEVICE_LIST sheet and return validated device records.
    """
    if warnings is None:
        warnings = []
    if errors is None:
        errors = []

    headers = [cell.value for cell in ws[1]]

    # Check for missing required columns
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        msg = f"Missing required columns: {', '.join(sorted(missing))}"
        if strict:
            raise RuntimeError(msg)
        errors.append(msg)
        return []

    devices = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = dict(zip(headers, row))
        record["_row_idx"] = row_idx  # ← single source of truth

        # Skip entirely blank rows silently
        if not any(v for k, v in record.items() if k != "_row_idx"):
            continue

        row_valid = True

        # Required fields
        for field in REQUIRED_COLUMNS:
            if not record.get(field):
                msg = f"Row {row_idx}: missing required value '{field}' - failed to create tags"
                if strict:
                    raise RuntimeError(msg)
                errors.append(msg)
                row_valid = False
                break

        # Optional fields — stored on the record so they can be filtered before reporting
        record["_warnings"] = [
            f"Row {row_idx}: missing optional value '{field}'"
            for field in OPTIONAL_COLUMNS
            if not record.get(field)
        ]

        if row_valid:
            devices.append(record)

    return devices


def is_control_row(row):
    """
    Determine if a row is a control row (row starting with ":").

    Args:
        row: A row from the Excel sheet.

    Returns:
        bool: True if it's a control row, False otherwise.
    """
    return bool(row) and isinstance(row[0], str) and row[0].startswith(":")


def expand_template(ws, device):
    """
    Expand template rows by replacing placeholders with actual values from the device dictionary.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet containing the template.
        device (dict): A device record with its details.

    Returns:
        list: List of expanded rows.
    """
    expanded_rows = []

    # Build base replacement map from PLACEHOLDER_MAP
    replacements = {
        token: str(device.get(column, "") or "")
        for token, column in PLACEHOLDER_MAP.items()
    }

    # Scan template for OFFSET+N tokens and pre-compute their values
    offset_raw = device.get("OFFSET")
    if offset_raw is not None:
        try:
            base_offset = int(offset_raw)
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, str):
                        for m in re.finditer(r'\{OFFSET\+(\d+)\}', cell):
                            n = int(m.group(1))
                            replacements[f'{{OFFSET+{n}}}'] = str(
                                base_offset + n)
        except (ValueError, TypeError):
            pass

    # Sort longest tokens first to prevent OFFSET from matching inside OFFSET+N
    sorted_tokens = sorted(replacements, key=len, reverse=True)

    for row in ws.iter_rows(values_only=True):
        new_row = []
        for cell in list(row):
            if isinstance(cell, str):
                for token in sorted_tokens:
                    cell = cell.replace(token, replacements[token])
            new_row.append(cell)
        expanded_rows.append(new_row)

    return expanded_rows


def check_tag_length(wb, limit):
    """
    For each template sheet, find the minimum remaining characters available
    for the HMI_TAG value given the fixed characters surrounding it in column A.

    Returns a list of (sheet_name, max_length) tuples, sorted by sheet name.
    Sheets with no qualifying rows are omitted.
    """
    results = []
    for name in list_templates(wb):
        ws = wb[name]
        min_remaining = None
        for row in ws.iter_rows(values_only=True):
            cell = row[0] if row else None
            if not isinstance(cell, str):
                continue
            if is_control_row(row):
                continue
            if "HMI_TAG" not in cell:
                continue
            fixed = len(cell) - len("HMI_TAG")
            remaining = limit - fixed
            if min_remaining is None or remaining < min_remaining:
                min_remaining = remaining
        if min_remaining is not None:
            results.append((name, min_remaining))
    return results


def list_columns(ws):
    """
    Inspect the DEVICE_LIST sheet and return required/optional columns found.
    """
    headers = [cell.value for cell in ws[1] if cell.value]

    found = set(headers)

    required_found = sorted(REQUIRED_COLUMNS & found)
    required_missing = sorted(REQUIRED_COLUMNS - found)

    optional_found = sorted(set(OPTIONAL_COLUMNS) & found)
    optional_missing = sorted(set(OPTIONAL_COLUMNS) - found)

    extra_columns = sorted(found - REQUIRED_COLUMNS - set(OPTIONAL_COLUMNS))

    return {
        "required_found": required_found,
        "required_missing": required_missing,
        "optional_found": optional_found,
        "optional_missing": optional_missing,
        "extra_columns": extra_columns,
        "all_columns": sorted(found),
    }


# ── UDT import helpers ────────────────────────────────────────────────────────

_IODISC_HEADER = [
    ':IODisc', 'Group', 'Comment', 'Logged', 'EventLogged', 'EventLoggingPriority',
    'RetentiveValue', 'InitialDisc', 'OffMsg', 'OnMsg', 'AlarmState', 'AlarmPri',
    'DConversion', 'AccessName', 'ItemUseTagname', 'ItemName', 'ReadOnly',
    'AlarmComment', 'AlarmAckModel', 'DSCAlarmDisable', 'DSCAlarmInhibitor', 'SymbolicName',
]

_IOREAL_HEADER = [
    ':IOReal', 'Group', 'Comment', 'Logged', 'EventLogged', 'EventLoggingPriority',
    'RetentiveValue', 'RetentiveAlarmParameters', 'AlarmValueDeadband', 'AlarmDevDeadband',
    'EngUnits', 'InitialValue', 'MinEU', 'MaxEU', 'Deadband', 'LogDeadband',
    'LoLoAlarmState', 'LoLoAlarmValue', 'LoLoAlarmPri',
    'LoAlarmState', 'LoAlarmValue', 'LoAlarmPri',
    'HiAlarmState', 'HiAlarmValue', 'HiAlarmPri',
    'HiHiAlarmState', 'HiHiAlarmValue', 'HiHiAlarmPri',
    'MinorDevAlarmState', 'MinorDevAlarmValue', 'MinorDevAlarmPri',
    'MajorDevAlarmState', 'MajorDevAlarmValue', 'MajorDevAlarmPri',
    'DevTarget', 'ROCAlarmState', 'ROCAlarmValue', 'ROCAlarmPri', 'ROCTimeBase',
    'MinRaw', 'MaxRaw', 'Conversion',
    'AccessName', 'ItemUseTagname', 'ItemName', 'ReadOnly', 'AlarmComment', 'AlarmAckModel',
    'LoLoAlarmDisable', 'LoAlarmDisable', 'HiAlarmDisable', 'HiHiAlarmDisable',
    'MinDevAlarmDisable', 'MajDevAlarmDisable', 'RocAlarmDisable',
    'LoLoAlarmInhibitor', 'LoAlarmInhibitor', 'HiAlarmInhibitor', 'HiHiAlarmInhibitor',
    'MinDevAlarmInhibitor', 'MajDevAlarmInhibitor', 'RocAlarmInhibitor',
    'SymbolicName', None,
]

_IOINT_HEADER = [':IOInt'] + _IOREAL_HEADER[1:]

_UDT_TYPE_HEADERS = {
    'BIT': _IODISC_HEADER,
    'REAL': _IOREAL_HEADER,
    'INT': _IOINT_HEADER,
    'DINT': _IOINT_HEADER,
}

_UDT_SKIP_TYPES = {'SINT', 'TIMER'}


def _udt_disc_row(tagname, item_name, comment):
    return [
        tagname, 'ALARM_GROUP', comment,
        'No', 'No', 0, 'No',
        'Off', 'Off', 'On', 'None', 1, 'Direct',
        'ACCESS_NAME', 'No', item_name, 'Yes',
        None, 0, 0, None, None,
    ]


def _udt_real_row(tagname, item_name, comment):
    return [
        tagname, 'ALARM_GROUP', comment,
        'No', 'No', 0, 'No', 'No', 0, 0,
        None, 0, 0, 100, 0, 0.1,
        'Off', 0, 1, 'Off', 0, 1,
        'Off', 0, 1, 'Off', 0, 1,
        'Off', 0, 1, 'Off', 0, 1,
        0, 'Off', 0, 1, 'Min', 0, 100, 'Linear',
        'ACCESS_NAME', 'No', item_name, 'Yes',
        None, 0,
        0, 0, 0, 0, 0, 0, 0,
        None, None, None, None, None, None, None,
        None, None,
    ]


def _udt_int_row(tagname, item_name, comment):
    return [
        tagname, 'ALARM_GROUP', comment,
        'No', 'No', 0, 'No', 'No', 0, 0,
        None, 0, -32768, 32767, 0, 0,
        'Off', 0, 1, 'Off', 0, 1,
        'Off', 0, 1, 'Off', 0, 1,
        'Off', 0, 1, 'Off', 0, 1,
        0, 'Off', 0, 1, 'Min', -32768, 32767, 'Linear',
        'ACCESS_NAME', 'No', item_name, 'Yes',
        None, 0,
        0, 0, 0, 0, 0, 0, 0,
        None, None, None, None, None, None, None,
        None, None,
    ]


_UDT_ROW_BUILDERS = {
    'BIT': _udt_disc_row,
    'REAL': _udt_real_row,
    'INT': _udt_int_row,
    'DINT': _udt_int_row,
}


def parse_udt_l5x(path):
    """
    Parse a Studio5000 .L5X UDT export file.
    Returns (udt_name, members, warnings) where members is a list of
    (member_name, datatype, description).  SINT and TIMER members are excluded.
    """
    try:
        tree = ET.parse(path)
    except ET.ParseError as e:
        raise RuntimeError(f"Failed to parse L5X file: {e}") from e

    root = tree.getroot()
    datatype_elem = root.find('.//DataType[@Use="Target"]')
    if datatype_elem is None:
        raise RuntimeError("No DataType[@Use='Target'] element found in L5X file")

    udt_name = datatype_elem.get('Name')
    members = []
    warnings = []

    for member in datatype_elem.findall('Members/Member'):
        dtype = member.get('DataType', '')
        if dtype in _UDT_SKIP_TYPES:
            continue

        name = member.get('Name', '')
        desc_elem = member.find('Description')
        if desc_elem is not None and desc_elem.text:
            desc = ' '.join(desc_elem.text.strip().split())
        else:
            desc = ''

        if dtype not in _UDT_TYPE_HEADERS:
            warnings.append(f"Member '{name}' has unsupported DataType '{dtype}' — skipped")
            continue

        members.append((name, dtype, desc))

    return udt_name, members, warnings


def import_udt(l5x_path, wb):
    """
    Parse a Studio5000 UDT .L5X file and add a template sheet to wb.
    Returns (udt_name, member_count, warnings).
    Raises RuntimeError if the sheet name already exists in the workbook.
    """
    udt_name, members, warnings = parse_udt_l5x(l5x_path)

    if udt_name in wb.sheetnames:
        raise RuntimeError(
            f"Sheet '{udt_name}' already exists in the workbook"
        )

    ws = wb.create_sheet(udt_name)

    prev_dtype = None
    for name, dtype, desc in members:
        if dtype != prev_dtype:
            ws.append(_UDT_TYPE_HEADERS[dtype])
            prev_dtype = dtype
        tagname = 'HMI_TAG_' + name.lstrip('_')
        item_name = 'PLC_TAG.' + name
        ws.append(_UDT_ROW_BUILDERS[dtype](tagname, item_name, desc))

    return udt_name, len(members), warnings


# ── end UDT import helpers ─────────────────────────────────────────────────────


def main():
    """
    Main function to generate the tag import CSV.
    """
    args = parse_args()
    warnings = []
    errors = []

    # Load the workbook
    try:
        wb = load_workbook(args.workbook, data_only=True)
    except FileNotFoundError:
        print(f"Error: file not found: {args.workbook!r}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Error: could not open workbook {args.workbook!r}: {e}", file=sys.stderr)
        sys.exit(1)

    # Handle --import-udt and exit early
    if args.import_udt:
        try:
            udt_name, count, udt_warnings = import_udt(args.import_udt, wb)
        except (RuntimeError, FileNotFoundError) as e:
            print(f"Error: {e}", file=sys.stderr)
            sys.exit(1)
        wb.save(args.workbook)
        print(f"Added template sheet '{udt_name}' ({count} tags) -> {args.workbook}")
        for w in udt_warnings:
            print(f"  Warning: {w}", file=sys.stderr)
        return

    # Handle --list-templates and exit early
    if args.list_templates:
        templates = list_templates(wb)

        if not templates:
            print("No templates found (only DEVICE_LIST present).")
        else:
            print(f"Found {len(templates)} template(s):\n")
            for name in templates:
                print(f"  {name}")

        return

    # Handle --list-columns and exit early
    if args.list_columns:
        if "DEVICE_LIST" not in wb.sheetnames:
            raise RuntimeError("DEVICE_LIST sheet not found")

        ws = wb["DEVICE_LIST"]
        info = list_columns(ws)

        print(f"DEVICE_LIST columns found ({len(info['all_columns'])}):\n")

        print("  REQUIRED:")
        if info["required_found"]:
            for col in info["required_found"]:
                print(f"    {col}")
        else:
            print("    (none)")

        if info["required_missing"]:
            print("\n  MISSING REQUIRED:")
            for col in info["required_missing"]:
                print(f"    {col}")

        print("\n  OPTIONAL:")
        if info["optional_found"]:
            for col in info["optional_found"]:
                print(f"    {col}")
        else:
            print("    (none)")

        if info["optional_missing"]:
            print("\n  OPTIONAL NOT PRESENT:")
            for col in info["optional_missing"]:
                print(f"    {col}")

        if info["extra_columns"]:
            print("\n  EXTRA COLUMNS:")
            for col in info["extra_columns"]:
                print(f"    {col}")

        return

    # Handle --tag-length and exit early
    if args.tag_length:
        if args.ww_version:
            limit = 128 if args.ww_version == "2023" else 32
        else:
            answer = input("Are you using Wonderware 2023 or later? [y/N]: ").strip().lower()
            limit = 128 if answer in ("y", "yes") else 32
        results = check_tag_length(wb, limit)
        if not results:
            print("No template sheets with HMI_TAG rows found.")
        else:
            print(f"Max HMI_TAG value length per template (Wonderware {'>= 2023' if limit == 128 else '<= 2020'}, limit = {limit} chars):\n")
            for name, max_len in results:
                print(f"  {name}: max HMI_TAG value = {max_len} chars")
        return

    # Check if the DEVICE_LIST sheet is present
    if "DEVICE_LIST" not in wb.sheetnames:
        raise RuntimeError("DEVICE_LIST sheet not found")

    device_ws = wb["DEVICE_LIST"]
    devices = read_device_list(
        device_ws,
        strict=args.strict,
        warnings=warnings,
        errors=errors,
    )

    # Apply --filter if specified
    if args.filter:
        col, _, val = args.filter.partition("=")
        if not col or not val:
            print(
                f"Error: --filter must be in COL=VAL format, got: {args.filter!r}", file=sys.stderr)
            sys.exit(1)
        devices = [d for d in devices if str(d.get(col, "")) == val]

    # Promote per-row warnings only for devices that survived filtering
    for device in devices:
        warnings.extend(device.pop("_warnings", []))

    output_rows = []
    tag_count = 0

    # Iterate over validated devices
    for device in devices:
        row_idx = device["_row_idx"]
        sheet_name = device["DEVICE_TYPE"]

        # Template existence check (same behavior as missing required value)
        if sheet_name not in wb.sheetnames:
            msg = f"Row {row_idx}: template sheet '{sheet_name}' not found - failed to create tags"
            if args.strict:
                raise RuntimeError(msg)
            errors.append(msg)
            continue

        template_ws = wb[sheet_name]
        rows = expand_template(template_ws, device)
        output_rows.extend(rows)

        for row in rows:
            if row and not is_control_row(row):
                tag_count += 1

    # Perform dry run or write the output file
    if args.dry_run:
        print("Dry run enabled — no output file written")
    else:
        with open(args.output, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([":mode=replace"])
            for row in output_rows:
                # Replace None values with empty strings
                writer.writerow(["" if v is None else v for v in row])

    # Print summary based on whether it's a dry run or not
    if args.dry_run:
        print(f"[DRY RUN] {tag_count} tags would be generated")
    else:
        print(f"Generated {tag_count} tags -> {args.output}")

    # ---- Validation Summary ----
    if warnings or errors:
        print("\nValidation Summary", file=sys.stderr)

        if errors:
            print(f"  Errors ({len(errors)}):", file=sys.stderr)
            for msg in errors:
                print(f"    - {msg}", file=sys.stderr)

        if warnings:
            print(f"  Warnings ({len(warnings)}):", file=sys.stderr)
            for msg in warnings:
                print(f"    - {msg}", file=sys.stderr)
    else:
        print("\nValidation Summary: no issues found")

    if errors:
        sys.exit(1)


if __name__ == "__main__":
    main()
